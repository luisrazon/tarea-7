# AUTORES
# MIGUEL ANGEL CISNEROS DAVILA
# JUAN ANGEL FACUNDO LOPEz
# JOSE LUIS RAZON LEYVA
# EDGAR OMAR SANDOVAL SALAZAR
from pyhunter import PyHunter
from openpyxl import Workbook
from openpyxl import load_workbook
import getpass
import os


def busqueda(organizacion):
    # Cantidad de resultados esperados de la búsqueda
    # El límite MENSUAL de Hunter es 50, cuidado!
    resultado = hunter.domain_search(company=organizacion, limit=1,
                                     emails_type='personal')
    return resultado


def informacion_guardada(datos_encontrados, organizacion):
    # Carga archivo excel, si no existe, lo crea
    try:
        libro = load_workbook("Hunter" + organizacion + ".xlsx")
    except FileNotFoundError:
        libro = Workbook()
    hoja = libro.create_sheet(organizacion)
    # Guarda archivo creado con nombre especificado
    libro.save("Hunter" + organizacion + ".xlsx")
    # Seleccionar hoja
    if len(libro.sheetnames) > 1:
        libro.active = 1
        hoja = libro.active
    else:
        hoja = libro.active
    # Encabezado de columnas
    hoja["A1"] = "Dominio"
    hoja["B1"] = "Patrón"
    hoja["C1"] = "Emails"
    hoja["D1"] = "Tipo de email"
    # Guardar datos usando llaves del resultado
    hoja["A2"] = datos_encontrados["domain"]
    hoja["B2"] = datos_encontrados["pattern"]
    # Guarda la lista de la llave "emails" en una nueva variable
    emails = datos_encontrados["emails"]
    # Se declara nueva variable donde se guarda el diccionario
    # del indice 0 de la lista emails
    dic_emails = emails[0]
    hoja["C2"] = dic_emails["value"]
    hoja["D2"] = dic_emails["type"]
    libro.save("Hunter" + organizacion + ".xlsx")

os.system("cls")
print("----------Script para buscar información----------")
apikey = getpass.getpass("Ingresa tu API key: ")
hunter = PyHunter(apikey)
orga = input("Dominio a investigar: ")
datos_encontrados = busqueda(orga)
if datos_encontrados is None:
    exit()
else:
    print(datos_encontrados)
    print(type(datos_encontrados))
    informacion_guardada(datos_encontrados, orga)
